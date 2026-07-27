import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from automation.config.config import Config
from automation.utils.logger import logger
from typing import List, Dict

class ExcelReporter:
    @staticmethod
    def generate_excel_reports(test_results: List[Dict]):
        """Generate all required Excel reports with multiple sheets and custom styling."""
        df = pd.DataFrame(test_results)
        if df.empty:
            logger.warning("No test results available to generate Excel reports.")
            return

        excel_path = Config.EXCEL_DIR / "Automation_Test_Report.xlsx"
        failed_path = Config.EXCEL_DIR / "Failed_Test_Cases.xlsx"
        passed_path = Config.EXCEL_DIR / "Passed_Test_Cases.xlsx"
        summary_path = Config.EXCEL_DIR / "Summary_Report.xlsx"

        # Sheet 1: Executed Test Cases
        # Sheet 2: Passed Tests
        # Sheet 3: Failed Tests
        # Sheet 4: Skipped Tests
        # Sheet 5: Execution Metrics
        # Sheet 6: Defect Summary

        passed_df = df[df["status"] == "PASSED"]
        failed_df = df[df["status"] == "FAILED"]
        skipped_df = df[df["status"] == "SKIPPED"]

        # Calculate metrics
        total = len(df)
        passed_cnt = len(passed_df)
        failed_cnt = len(failed_df)
        skipped_cnt = len(skipped_df)
        pass_rate = round((passed_cnt / total * 100), 2) if total > 0 else 0

        metrics_df = pd.DataFrame([
            {"Metric": "Total Test Cases", "Value": total},
            {"Metric": "Passed Test Cases", "Value": passed_cnt},
            {"Metric": "Failed Test Cases", "Value": failed_cnt},
            {"Metric": "Skipped Test Cases", "Value": skipped_cnt},
            {"Metric": "Pass Rate (%)", "Value": f"{pass_rate}%"},
            {"Metric": "Target Deployed URL", "Value": Config.BASE_URL}
        ])

        defect_df = failed_df[["test_id", "module", "test_name", "failure_reason"]] if not failed_df.empty else pd.DataFrame(columns=["test_id", "module", "test_name", "failure_reason"])

        # Write to main Excel workbook
        with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="Executed Test Cases", index=False)
            passed_df.to_excel(writer, sheet_name="Passed Tests", index=False)
            failed_df.to_excel(writer, sheet_name="Failed Tests", index=False)
            skipped_df.to_excel(writer, sheet_name="Skipped Tests", index=False)
            metrics_df.to_excel(writer, sheet_name="Execution Metrics", index=False)
            defect_df.to_excel(writer, sheet_name="Defect Summary", index=False)

        # Style main excel
        ExcelReporter._apply_excel_styles(excel_path)

        # Write auxiliary single-purpose excel reports
        with pd.ExcelWriter(failed_path, engine="openpyxl") as writer:
            failed_df.to_excel(writer, sheet_name="Failed Tests", index=False)
        ExcelReporter._apply_excel_styles(failed_path)

        with pd.ExcelWriter(passed_path, engine="openpyxl") as writer:
            passed_df.to_excel(writer, sheet_name="Passed Tests", index=False)
        ExcelReporter._apply_excel_styles(passed_path)

        with pd.ExcelWriter(summary_path, engine="openpyxl") as writer:
            metrics_df.to_excel(writer, sheet_name="Execution Metrics", index=False)
        ExcelReporter._apply_excel_styles(summary_path)

        logger.info(f"Excel reports generated successfully in {Config.EXCEL_DIR}")

    @staticmethod
    def _apply_excel_styles(filepath):
        wb = openpyxl.load_workbook(filepath)
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        
        pass_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        fail_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )

        for sheetname in wb.sheetnames:
            ws = wb[sheetname]
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            for row in range(2, ws.max_row + 1):
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.border = thin_border
                    val_str = str(cell.value).upper()
                    if val_str == "PASSED":
                        cell.fill = pass_fill
                    elif val_str == "FAILED":
                        cell.fill = fail_fill

            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

        wb.save(filepath)
