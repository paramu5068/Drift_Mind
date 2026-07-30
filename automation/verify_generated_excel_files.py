import openpyxl
import os

files = [
    "Automation_Test_Report.xlsx",
    "Unit_Test_Cases.xlsx",
    "vulnerability_test_report.xlsx",
    "Validation_Test_Cases.xlsx",
    "Deploy_Test_Cases.xlsx",
    "Load_Test_Cases.xlsx",
    "Passed_Test_Cases.xlsx"
]

base_dir = r"c:\Users\nasri\OneDrive\Desktop\drift_mind"

print("--- VERIFICATION REPORT ---")
for f in files:
    file_path = os.path.join(base_dir, f)
    wb = openpyxl.load_workbook(file_path)
    sheets = wb.sheetnames
    ws_detail = wb["Detailed Test Cases"]
    row_count = ws_detail.max_row - 1 # excluding header
    
    # Check for placeholder strings in first 50 rows
    placeholders = 0
    for r in range(2, ws_detail.max_row + 1):
        step_val = str(ws_detail.cell(row=r, column=5).value)
        if "test case step " in step_val.lower() or "check i" in step_val.lower():
            placeholders += 1

    print(f"File: {f}")
    print(f"  Sheets: {sheets}")
    print(f"  Total Detailed Test Cases: {row_count}")
    print(f"  Placeholder test step occurrences: {placeholders}")
    print(f"  Sample Row 2 Test Name: {ws_detail.cell(row=2, column=3).value}")
    print(f"  Sample Row 2 Precondition: {ws_detail.cell(row=2, column=4).value}")
    print(f"  Sample Row 2 Test Step: {ws_detail.cell(row=2, column=5).value}")
    print(f"  Sample Row 2 Expected Result: {ws_detail.cell(row=2, column=6).value}")
    print("-" * 50)
